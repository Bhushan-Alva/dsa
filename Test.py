import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle


def compare_and_export_amount_diff(
    current_df,
    previous_df,
    key_col="Project Number",
    value_col="Approved Amount",
    diff_col="Amount_Diff",
    output_file="project_amount_diff_check.xlsx",
    sheet_name="Data",
    negative_color="F8D7DA",
    number_format="0.00%",
    max_col_width=40
):
    """
    Compare two datasets by aggregating a value column and exporting % difference to Excel

    Workflow:
    - Group + sum both datasets
    - Outer merge
    - Fill missing values
    - Calculate % diff
    - Sort
    - Export to Excel
    - Format + highlight negative diffs
    """

    # ------------------ 1) Aggregate ------------------
    cur_grp = (
        current_df
        .groupby(key_col, as_index=False)
        .agg({value_col: "sum"})
        .rename(columns={value_col: f"{value_col}_new"})
    )

    prev_grp = (
        previous_df
        .groupby(key_col, as_index=False)
        .agg({value_col: "sum"})
        .rename(columns={value_col: f"{value_col}_old"})
    )

    # ------------------ 2) Merge ------------------
    result = cur_grp.merge(
        prev_grp,
        on=key_col,
        how="outer"
    )

    # ------------------ 3) Clean ------------------
    result[[f"{value_col}_new", f"{value_col}_old"]] = (
        result[[f"{value_col}_new", f"{value_col}_old"]].fillna(0)
    )

    # ------------------ 4) Diff Calculation ------------------
    result[diff_col] = np.where(
        result[f"{value_col}_old"] == 0,
        1,  # 100% when previous is zero
        result[f"{value_col}_new"].div(result[f"{value_col}_old"]) - 1
    )

    # ------------------ 5) Sort ------------------
    result = result.sort_values(diff_col)

    # ------------------ 6) Export ------------------
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        result.to_excel(writer, index=False, sheet_name=sheet_name)

    # ------------------ 7) Excel Formatting ------------------
    wb = load_workbook(output_file)
    ws = wb[sheet_name]

    # Find diff column dynamically
    headers = [cell.value for cell in ws[1]]
    if diff_col not in headers:
        raise ValueError(f"Column '{diff_col}' not found in Excel sheet")

    col_idx = headers.index(diff_col) + 1
    col_letter = ws.cell(row=1, column=col_idx).column_letter

    max_row = ws.max_row
    diff_range = f"{col_letter}2:{col_letter}{max_row}"

    # Number format
    for row in range(2, max_row + 1):
        ws[f"{col_letter}{row}"].number_format = number_format

    # Conditional formatting (negative values)
    fill = PatternFill(
        start_color=negative_color,
        end_color=negative_color,
        fill_type="solid"
    )
    dxf = DifferentialStyle(fill=fill)

    rule = Rule(
        type="expression",
        dxf=dxf,
        formula=[f"{col_letter}2<0"]
    )

    ws.conditional_formatting.add(diff_range, rule)

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = min(max_length + 2, max_col_width)

    # Save
    wb.save(output_file)

    print(f"Saved: {output_file}")

    return result
        
result = compare_and_export_amount_diff(
    current_df=data,
    previous_df=prev_data,
    key_col="Project Number",
    value_col="Approved Amount",
    output_file="project_amount_diff_check.xlsx"
)




# ================================
# FULL REUSABLE REPORTING ENGINE
# ================================

import numpy as np
import pandas as pd
import win32com.client as win32


# -------------------------------------------------
# 1) BUILD COUNTRY x MONTH % DIFFERENCE MATRIX
# -------------------------------------------------
def build_pct_diff_matrix(
    latest_raw: pd.DataFrame,
    old_raw: pd.DataFrame,
    country_col="Employee Country",
    date_col="Report Extract Date",
    value_col="Approved Amount"
):
    """
    Builds Country x Month % Difference Matrix
    Returns:
        matrix_view (DataFrame)
        latest_month_name (str)
    """

    month_order = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

    latest = latest_raw.copy()
    old = old_raw.copy()

    latest["Month"] = latest[date_col].dt.strftime("%b")
    old["Month"] = old[date_col].dt.strftime("%b")

    latest_months = pd.Series(latest["Month"]).dropna().astype(str).tolist()
    valid_latest_months = [m for m in month_order if m in set(latest_months)]
    latest_month_name = valid_latest_months[-1] if valid_latest_months else None

    latest_grp = (
        latest
        .groupby([country_col, "Month"], as_index=False)[value_col]
        .sum()
    )

    old_grp = (
        old
        .groupby([country_col, "Month"], as_index=False)[value_col]
        .sum()
    )

    comparison = (
        pd.merge(
            latest_grp,
            old_grp,
            on=[country_col, "Month"],
            how="outer",
            suffixes=("_latest", "_old")
        )
        .fillna(0.0)
    )

    latest_vals = comparison[f"{value_col}_latest"]
    old_vals = comparison[f"{value_col}_old"]

    comparison["Pct_Difference"] = np.where(
        old_vals == 0,
        np.where(latest_vals == 0, 0.0, np.sign(latest_vals) * 100.0),
        ((latest_vals - old_vals) / old_vals) * 100.0
    ).round(2)

    matrix_view = (
        comparison
        .pivot_table(
            index=country_col,
            columns="Month",
            values="Pct_Difference",
            aggfunc="first",
            fill_value=0.0
        )
        .reindex(columns=month_order)
    )

    return matrix_view, latest_month_name


# -------------------------------------------------
# 2) RENDER OUTLOOK-SAFE HTML TABLE
# -------------------------------------------------
def render_html_matrix(df: pd.DataFrame, latest_month_name: str) -> str:
    """
    Renders an HTML table with color rules:
    Latest month:
        0.00 -> RED
        non-zero -> BLUE
    Other months:
        0.00 -> GREEN
        non-zero -> RED
    """

    RED = "#FFD5D5"
    BLUE = "#CFE8FF"
    GREEN = "#C6EFCE"

    table_style = (
        "border-collapse: collapse; "
        "font-family: Segoe UI, Roboto, Arial; "
        "font-size: 12px;"
    )

    cell_style = "border: 1px solid #bbb; padding: 6px 8px; text-align: center;"
    row_header_style = cell_style + " text-align: left; font-weight: bold;"

    html = [f'<table style="{table_style}">']

    # Header
    html.append("<thead><tr>")
    html.append(f'<th style="{cell_style}"></th>')
    for col in df.columns:
        html.append(f'<th style="{cell_style}">{col}</th>')
    html.append("</tr></thead>")

    # Body
    html.append("<tbody>")
    for idx, row in df.iterrows():
        html.append("<tr>")
        html.append(f'<th style="{row_header_style}">{idx}</th>')

        for col in df.columns:
            val = row[col]

            if pd.isna(val):
                disp = "NA"
                style = cell_style
            else:
                val = float(val)
                disp = f"{val:.2f}"
                is_zero = abs(val) < 1e-12

                if latest_month_name and col == latest_month_name:
                    bg = RED if is_zero else BLUE
                else:
                    bg = GREEN if is_zero else RED

                style = f"{cell_style} background-color: {bg};"

            html.append(f'<td style="{style}">{disp}</td>')

        html.append("</tr>")
    html.append("</tbody></table>")

    return "\n".join(html)


# -------------------------------------------------
# 3) SEND EMAIL VIA OUTLOOK
# -------------------------------------------------
def send_outlook_email(
    to_emails,
    subject,
    message,
    html_table,
    excel_path=None,
    cc_emails=None,
    bcc_emails=None,
    preview=False
):
    """
    Sends an HTML email via Outlook with optional attachment
    """

    legend_html = """
    <div style="margin-bottom:8px;">
        <span style="display:inline-block;width:14px;height:14px;background:#FFD5D5;border:1px solid #bbb;"></span>
        <span style="margin-right:18px;">RED: non-zero (other months) or zero (latest month)</span>
        <span style="display:inline-block;width:14px;height:14px;background:#CFE8FF;border:1px solid #bbb;"></span>
        <span style="margin-right:18px;">BLUE: non-zero in latest month</span>
        <span style="display:inline-block;width:14px;height:14px;background:#C6EFCE;border:1px solid #bbb;"></span>
        <span>GREEN: zero in other months</span>
    </div>
    """

    html_body = f"""
    <html>
    <body style="font-family: Segoe UI, Roboto, Arial; font-size: 14px;">
        <pre style="white-space: pre-wrap; margin: 0 0 12px 0;">{message.strip()}</pre>
        {legend_html}
        {html_table}
    </body>
    </html>
    """

    outlook = win32.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)

    mail.To = "; ".join(to_emails)

    if cc_emails:
        mail.CC = "; ".join(cc_emails)

    if bcc_emails:
        mail.BCC = "; ".join(bcc_emails)

    mail.Subject = subject
    mail.HTMLBody = html_body

    if excel_path:
        mail.Attachments.Add(Source=excel_path)

    if preview:
        mail.Display()
    else:
        mail.Send()

    print("Email sent via Outlook!")


# -------------------------------------------------
# 4) ONE-LINE PIPELINE RUNNER
# -------------------------------------------------
def run_country_month_report(
    latest_raw,
    old_raw,
    to_emails,
    subject,
    custom_message,
    excel_path=None,
    cc_emails=None,
    preview=True
):
    matrix, latest_month = build_pct_diff_matrix(
        latest_raw=latest_raw,
        old_raw=old_raw
    )

    html_table = render_html_matrix(matrix, latest_month)

    send_outlook_email(
        to_emails=to_emails,
        cc_emails=cc_emails,
        subject=subject,
        message=custom_message,
        html_table=html_table,
        excel_path=excel_path,
        preview=preview
    )

    return matrix


# -------------------------------------------------
# 5) EXAMPLE USAGE
# -------------------------------------------------
# matrix = run_country_month_report(
#     latest_raw=latest_raw_df,
#     old_raw=old_raw_df,
#     to_emails=["bhushan.alva@capgemini.com"],
#     cc_emails=["cat_myexpense.in@capgemini.com"],
#     subject="MyExpense Company Car - Country x Month % Difference",
#     custom_message="""
# Hello All,
#
# This is the automated MyExpense report for latest vs previous month.
# Please find the % difference matrix below.
#
# Thanks,
# """,
#     excel_path="Difference.xlsx",
#     preview=True   # Set False to auto-send
# )
